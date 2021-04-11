import openpyxl
from openpyxl import Workbook
import pathlib

from tkinter import *
from tkinter import ttk
from tkinter import filedialog

from PIL import ImageTk,Image
import os

#EXCEL WORKBOOK
wb = Workbook ()
wb = openpyxl.load_workbook (r'URSA Database.xlsx', data_only = True)
ws = wb.worksheets[0]

#ROOT WINDOW START
root = Tk()
root.geometry("1300x650")
#root.configure (bg="white")
root.title('URSA')
root.iconbitmap (r'URSA.ico')

#Title Label:TKinter Widget
T1 = Label(root, text='URSA', font=('Industry Inc Base', 62, 'normal')).place(x=15, y=5)
T2 = Label(root, text='Urban Resilience Score Algorithm', font=('Industry Inc Base', 9, 'normal')).place(x=15, y=100)
T3 = Label(root, text='Welcome',font=('Industry Inc Base', 12, 'normal')).place(x=15, y=155)
T4 = Label(root, text='URSA is a planning tool striving for resilient development.', fg='#575757',font=('Myanmar Text', 8, 'normal')).place(x=15, y=180)
T5 = Label(root, text='URSAs algorithm consider 2 factors:', fg='#575757',font=('Myanmar Text', 8, 'normal')).place(x=15, y=198)
T6 = Label(root, text='Site Requirements', font=('Industry Inc Base', 8, 'normal')).place(x=15, y=230)
T7 =Label(root, text='Development Impact', font=('Industry Inc Base', 8, 'normal')).place(x=230, y=230)

#1)Site Requirements Void map
def open1():
    global my_img
    top1 = Toplevel()
    top1.title ('URSA: Map of Urban Voids, 1:125000 (M58)')
    top1.iconbitmap(r"URSA.ico")
    my_img = ImageTk.PhotoImage(Image.open(r"my_img.png"))
    my_label = Label(top1,image=my_img).pack()
    lambimg.config(image=my_img)

voidmapbutton = Button(root, width=18, height=1,text='Map of Urban Voids',fg='#282828',font=('Industry Inc Base', 8, 'normal'),bg='#CDC6C6', command=open1)
voidmapbutton.pack()
voidmapbutton.place (x=15, y=257)

#2)Development Impact Database
def open2():
    global my_img2
    top2 = Toplevel()
    top2.title ('URSA: Database of Resilience Determinants and Impact')
    top2.iconbitmap(r"URSA.ico")
    my_img2 = ImageTk.PhotoImage(Image.open(r"my_img2.png"))
    my_label2 = Label(top2,image=my_img2).pack()
    lambimg.config(image=my_img2)

voidmapbutton = Button(root, width=19, height=1,text='Resilience Database',fg='#282828',font=('Industry Inc Base', 8, 'normal'),bg='#CDC6C6', command=open2)
voidmapbutton.pack()
voidmapbutton.place (x=230, y=257)

Label(root, text='To generate an URSA Report, please complete the following:', fg='#575757',font=('Myanmar Text', 8, 'normal')).place(x=15, y=300)

#1. NAME OF DEVELOPMENT: User Input 1
L1 = Label(root, text='Development Name',fg='#3C0947',font=('Industry Inc Base', 12, 'normal')).place(x=15, y=340)

#Callback entry box
dn_inp = StringVar()
def dn_inp_callback():
    dn_inp.set(dn_E1.get())

dn_E1 = Entry(root, bd = 3, selectborderwidth=20)
dn_E1.pack(side = RIGHT)
dn_E1.place (x=15, y=370,width=260,height=30)
dn_E1.focus_set()
dn_B1 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 8, 'normal'), command=dn_inp_callback)
dn_B1.place(x=295, y=370)
dn_L1 = Label(root,textvariable=dn_inp,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=15, y=405)

#2. SITE REQUIREMENTS: User Input 2
L2 = Label(root, text='Site Requirements', fg='#3C0947',font=('Industry Inc Base', 12, 'normal')).place(x=15, y=430)
Label(root, text='Does the site lie within a void type? Enter Y/N for each.', fg='#575757',font=('Myanmar Text', 8, 'normal')).place(x=15, y=455)

#Widget List Box Void names
srlist = Listbox(root,width=32, height=6, fg='#6c6c6c',font=('Calibri', 11, 'normal'))

srlist.place (x=15,y=485)
srlist.insert(1, " 1. Void of Digital Connection")
srlist.insert(2, " 2. Void of Urban Connectivity")
srlist.insert(3, " 3. Void of Flood ")
srlist.insert(4, " 4. Void of Urban Vulnerability")
srlist.insert(5, " 5. Void of Shock Absorption")
srlist.insert(6, " 6. No Voids Present")
srlist.pack

#URSA ALGORITHM Variables
options = ["Void of Digital Connection", "Void of Urban Connectivity", "Void of Flood", "Voids of Urban Vulnerability", "Void of Shock Absorption","No Void"]
boo = ""
boo1 = ""
boo2 = ""
boo3 = ""
boo4 = ""
boo5 = ""
inp1 = ""
inp2 = ""
inp3 = ""
inp4 = ""
inp5 = ""
inp = ""
impact = ""
req = 1.5

# Print Base Site Requirements
def aerial_req():
    for i in range(4,20):
        ws.cell(row=84, column=i).value = ws.cell(row=5, column=i).value + ws.cell(row=9, column=i).value + ws.cell(row=13, column=i).value + ws.cell(row=19, column=i).value + ws.cell(row=22, column=i).value + ws.cell(row=24, column=i).value + ws.cell(row=29, column=i).value + ws.cell(row=31, column=i).value + ws.cell(row=34, column=i).value + ws.cell(row=37, column=i).value + ws.cell(row=42, column=i).value + ws.cell(row=47, column=i).value + ws.cell(row=51, column=i).value + ws.cell(row=59, column=i).value + ws.cell(row=62, column=i).value + ws.cell(row=67, column=i).value + ws.cell(row=72, column=i).value
aerial_req()

#(sr_inp1) 1. Voids of Digital Connection Y/N Callback
sr_inp1 = StringVar()
def sr_inp1_callback():
    sr_inp1.set(sr_E1.get())

sr_E1 = Entry(root, bd =3, selectborderwidth=5)
sr_E1.pack(side = RIGHT)
sr_E1.place (x=265, y=485,width=30,height=20)
sr_E1.focus_set()
sr_B1 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=sr_inp1_callback)
sr_B1.place(x=315, y=485)
sr_L1 = Label(root,textvariable=sr_inp1,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=380, y=485 )

#(sr_inp2) 2. Voids of Urban Connectivity Y/N
sr_inp2 = StringVar()
def sr_inp2_callback(): 
    sr_inp2.set(sr_E2.get())

sr_E2 = Entry(root, bd =3, selectborderwidth=5)
sr_E2.pack(side = RIGHT)
sr_E2.place (x=265, y=505,width=30,height=20)
sr_E2.focus_set()
sr_B2 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=sr_inp2_callback)
sr_B2.place(x=315, y=505)
sr_L2 = Label(root,textvariable=sr_inp2,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=380, y=505 )

#(sr_inp3) 3. Voids of Flood Y/N
sr_inp3 = StringVar()
def sr_inp3_callback(): 
    sr_inp3.set(sr_E3.get())

sr_E3 = Entry(root, bd =3, selectborderwidth=5)
sr_E3.pack(side = RIGHT)
sr_E3.place (x=265, y=525,width=30,height=20)
sr_E3.focus_set()
sr_B3 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=sr_inp3_callback)
sr_B3.place(x=315, y=525)
sr_L3 = Label(root,textvariable=sr_inp3,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=380, y=525 )

#(sr_inp4) 4. Voids of Urban Vulnerability Y/N
sr_inp4 = StringVar()
def sr_inp4_callback(): 
    sr_inp4.set(sr_E4.get())

sr_E4 = Entry(root, bd =3, selectborderwidth=5)
sr_E4.pack(side = RIGHT)
sr_E4.place (x=265, y=545,width=30,height=20)
sr_E4.focus_set()
sr_B4 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=sr_inp4_callback)
sr_B4.place(x=315, y=545)
sr_L4 = Label(root,textvariable=sr_inp4,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=380, y=545 )

#(sr_inp5) 5. Voids of Shock Absorption Y/N
sr_inp5 = StringVar()
def sr_inp5_callback(): 
    sr_inp5.set(sr_E5.get())

sr_E5 = Entry(root, bd =3, selectborderwidth=5)
sr_E5.pack(side = RIGHT)
sr_E5.place (x=265, y=565,width=30,height=20)
sr_E5.focus_set()
sr_B5 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=sr_inp5_callback)
sr_B5.place(x=315, y=565)
sr_L5 = Label(root,textvariable=sr_inp5,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=380, y=565 )

#(sr_inp6) 6. No Voids
sr_inp6 = StringVar()
def sr_inp6_callback(): 
    sr_inp6.set(sr_E6.get())

sr_E6 = Entry(root, bd =3, selectborderwidth=5)
sr_E6.pack(side = RIGHT)
sr_E6.place (x=265, y=585,width=30,height=20)
sr_E6.focus_set()
sr_B6 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=sr_inp6_callback)
sr_B6.place(x=315, y=585)
sr_L6 = Label(root,textvariable=sr_inp6,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=380, y=583 )

#3.DEVELOPMENT IMPACT:User Impact 3
L3 = Label(root, text='Development Impact', fg='#3C0947',font=('Industry Inc Base', 12, 'normal')).place(x=500, y=155)
Label(root, text='Rate your developments impact on each resilience factor below.', fg='#575757',font=('Myanmar Text', 8, 'normal')).place(x=500, y=180)
Label(root, text='Enter a value from 0-3 for each (0=No Impact, 3=High Impact)', fg='#575757',font=('Myanmar Text', 8, 'normal')).place(x=500, y=198)

#3.Button Determinants of Urban Resilience
di_L1=Label(root, text='Determinants of Urban Resilience', font=('Industry Inc Base', 8, 'normal')).place(x=500, y=230)

def open3():
    global my_img3
    top3 = Toplevel()
    top3.title ('URSA: The Determinants of Urban Resilience')
    top3.iconbitmap(r"URSA.ico")
    my_img3 = ImageTk.PhotoImage(Image.open(r"my_img3.png"))
    my_label3 = Label(top3,image=my_img3).pack()
    lambimg.config(image=my_img3)

determinants_B3 = Button(root, width=27, height=1,text='Pillars of Urban Resilience',fg='#282828',font=('Industry Inc Base', 8, 'normal'),bg='#CDC6C6', command=open3)
determinants_B3.pack()
determinants_B3.place (x=500, y=257)

#List Box: Infrastructure Pillar
di_list = Listbox(root,width=27, fg='#6c6c6c',height=16,font=('Calibri', 11, 'normal'))
di_list.place (x=500,y=300)
di_list.insert(1, " Water")
di_list.insert(2, " Energy")
di_list.insert(3, " Spatial Configuration")
di_list.insert(4, " Connectivity")
di_list.insert(5, " Transportation")
di_list.insert(6, " Green Infrastructure")
di_list.insert(7, " Defense Infrastructure")
di_list.insert(8, " Building and Design")
di_list.insert(9, " Technology and Information")
di_list.insert(10, " Surveillance")
di_list.insert(11, " Ecosystem")
di_list.insert(12, " Social & Demographic")
di_list.insert(13, " Health")
di_list.insert(14, " Planning")
di_list.insert(15, " Centralisation")
di_list.insert(16, " Economy")
di_list.pack

#(di_inp1) 1.Water
di_inp1 = StringVar()
def di_inp1_callback(): 
    di_inp1.set(di_E1.get())

di_E1 = Entry(root, bd =3, selectborderwidth=5)
di_E1.pack(side = RIGHT)
di_E1.place (x=716, y=300,width=30,height=20)
di_E1.focus_set()
di_B1 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp1_callback)
di_B1.place(x=763, y=300)
di_L1 = Label(root,textvariable=di_inp1,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=300 )

#(di_inp2) 2.Energy
di_inp2 = StringVar()
def di_inp2_callback(): 
    di_inp2.set(di_E2.get())

di_E2 = Entry(root, bd =3, selectborderwidth=5)
di_E2.pack(side = RIGHT)
di_E2.place (x=716, y=318,width=30,height=20)
di_E2.focus_set()
di_B2 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp2_callback)
di_B2.place(x=763, y=318)
di_L2 = Label(root,textvariable=di_inp2,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=318 )

#(di_inp3) 3.Spatial Configuration
di_inp3 = StringVar()
def di_inp3_callback(): 
    di_inp3.set(di_E3.get())

di_E3 = Entry(root, bd =3, selectborderwidth=5)
di_E3.pack(side = RIGHT)
di_E3.place (x=716, y=337,width=30,height=20)
di_E3.focus_set()
di_B3 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp3_callback)
di_B3.place(x=763, y=337)
di_L3 = Label(root,textvariable=di_inp3,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=337 )

#(di_inp4) 4.Connectivity
di_inp4 = StringVar()
def di_inp4_callback(): 
    di_inp4.set(di_E4.get())

di_E4 = Entry(root, bd =3, selectborderwidth=5)
di_E4.pack(side = RIGHT)
di_E4.place (x=716, y=356,width=30,height=20)
di_E4.focus_set()
di_B4 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp4_callback)
di_B4.place(x=763, y=356)
di_L4 = Label(root,textvariable=di_inp4,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=356 )

#(di_inp5) 5.Transportation
di_inp5 = StringVar()
def di_inp5_callback(): 
    di_inp5.set(di_E5.get())

di_E5 = Entry(root, bd =3, selectborderwidth=5)
di_E5.pack(side = RIGHT)
di_E5.place (x=716, y=376,width=30,height=20)
di_E5.focus_set()
di_B5 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp5_callback)
di_B5.place(x=763, y=376)
di_L5 = Label(root,textvariable=di_inp5,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=376 )

#(di_inp6) 6.Green Infrasturcture
di_inp6 = StringVar()
def di_inp6_callback(): 
    di_inp6.set(di_E6.get())

di_E6 = Entry(root, bd =3, selectborderwidth=5)
di_E6.pack(side = RIGHT)
di_E6.place (x=716, y=396,width=30,height=20)
di_E6.focus_set()
di_B6 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp6_callback)
di_B6.place(x=763, y=396)
di_L6 = Label(root,textvariable=di_inp6,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=396 )

#(di_inp7) 7.Defense Infrasturcture
di_inp7 = StringVar()
def di_inp7_callback(): 
    di_inp7.set(di_E7.get())

di_E7 = Entry(root, bd =3, selectborderwidth=5)
di_E7.pack(side = RIGHT)
di_E7.place (x=716, y=415,width=30,height=20)
di_E7.focus_set()
di_B7 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp7_callback)
di_B7.place(x=763, y=415)
di_L7 = Label(root,textvariable=di_inp7,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=415 )

#(di_inp8) 8.Building + Design
di_inp8 = StringVar()
def di_inp8_callback(): 
    di_inp8.set(di_E8.get())

di_E8 = Entry(root, bd =3, selectborderwidth=5)
di_E8.pack(side = RIGHT)
di_E8.place (x=716, y=433,width=30,height=20)
di_E8.focus_set()
di_B8 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp8_callback)
di_B8.place(x=763, y=433)
di_L8 = Label(root,textvariable=di_inp8,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=433 )

#(di_inp9) 9.Technology + Information
di_inp9 = StringVar()
def di_inp9_callback(): 
    di_inp9.set(di_E9.get())

di_E9 = Entry(root, bd =3, selectborderwidth=5)
di_E9.pack(side = RIGHT)
di_E9.place (x=716, y=451,width=30,height=20)
di_E9.focus_set()
di_B9 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp9_callback)
di_B9.place(x=763, y=451)
di_L9 = Label(root,textvariable=di_inp9,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=451 )

#(di_inp10) 10.Security
di_inp10 = StringVar()
def di_inp10_callback(): 
    di_inp10.set(di_E10.get())

di_E10 = Entry(root, bd =3, selectborderwidth=5)
di_E10.pack(side = RIGHT)
di_E10.place (x=716, y=470,width=30,height=20)
di_E10.focus_set()
di_B10 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp10_callback)
di_B10.place(x=763, y=470)
di_L10 = Label(root,textvariable=di_inp10,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=470 )

#(di_inp11) 11.Ecosystem
di_inp11 = StringVar()
def di_inp11_callback(): 
    di_inp11.set(di_E11.get())

di_E11 = Entry(root, bd =3, selectborderwidth=5)
di_E11.pack(side = RIGHT)
di_E11.place (x=716, y=489,width=30,height=20)
di_E11.focus_set()
di_B11 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp11_callback)
di_B11.place(x=763, y=489)
di_L11 = Label(root,textvariable=di_inp11,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=489 )

#(di_inp12) 12.Society
di_inp12 = StringVar()
def di_inp12_callback(): 
    di_inp12.set(di_E12.get())

di_E12 = Entry(root, bd =3, selectborderwidth=5)
di_E12.pack(side = RIGHT)
di_E12.place (x=716, y=509,width=30,height=20)
di_E12.focus_set()
di_B12 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp12_callback)
di_B12.place(x=763, y=509)
di_L12 = Label(root,textvariable=di_inp12,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=509 )

#(di_inp13) 13.Health
di_inp13 = StringVar()
def di_inp13_callback(): 
    di_inp13.set(di_E13.get())

di_E13 = Entry(root, bd =3, selectborderwidth=5)
di_E13.pack(side = RIGHT)
di_E13.place (x=716, y=529,width=30,height=20)
di_E13.focus_set()
di_B13 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp13_callback)
di_B13.place(x=763, y=529)
di_L13 = Label(root,textvariable=di_inp13,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=529 )

#(di_inp14) 14.Planning
di_inp14 = StringVar()
def di_inp14_callback(): 
    di_inp14.set(di_E14.get())

di_E14 = Entry(root, bd =3, selectborderwidth=5)
di_E14.pack(side = RIGHT)
di_E14.place (x=716, y=547,width=30,height=20)
di_E14.focus_set()
di_B14 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp14_callback)
di_B14.place(x=763, y=547)
di_L14 = Label(root,textvariable=di_inp14,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=547 )

#(di_inp15) 15.Centralisation
di_inp15 = StringVar()
def di_inp15_callback(): 
    di_inp15.set(di_E15.get())

di_E15 = Entry(root, bd =3, selectborderwidth=5)
di_E15.pack(side = RIGHT)
di_E15.place (x=716, y=567,width=30,height=20)
di_E15.focus_set()
di_B15 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp15_callback)
di_B15.place(x=763, y=567)
di_L15 = Label(root,textvariable=di_inp15,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=567 )

#(di_inp16) 16. Economy
di_inp16 = StringVar()
def di_inp16_callback(): 
    di_inp16.set(di_E16.get())

di_E16 = Entry(root, bd =3, selectborderwidth=5)
di_E16.pack(side = RIGHT)
di_E16.place (x=716, y=588,width=30,height=20)
di_E16.focus_set()
di_B16 = Button(root,text="Save", width=10,bg='gray', font=('Industry Inc Base', 6, 'normal'),command=di_inp16_callback)
di_B16.place(x=763, y=588)
di_L16 = Label(root,textvariable=di_inp16,font=('Industry Inc Base', 9, 'italic'),fg='#B00A11'). place(x=838, y=588 )

#PRINT EVERYTHING TO EXCEL
#Print Development Name
def printdn():
    ws.cell(row=94, column=2, value=dn_E1.get())
    wb.save(r'URSA Report.xlsx')

#Print Site Requirements
def printsitereq():
    boo1 = sr_E1.get()
    if boo1 == "Y":
        ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
        ws.cell(row=87, column=2).value = "Void of Digital Connection"
        ws.cell(row=87, column=12).value = "yes"
    else:
        ws.cell(row=87, column=2).value = ""
    boo2 = sr_E2.get()
    if boo2 == "Y":
        ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
        ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
        ws.cell(row=84, column=6).value = ws.cell(row=84, column=6).value * req
        ws.cell(row=88, column=2).value = "Void of Urban Connectivity"
        ws.cell(row=88, column=12).value = "yes"
        ws.cell(row=88, column=7).value = "yes"
        ws.cell(row=88, column=6).value = "yes"
    else:
        ws.cell(row=88, column=2).value = ""
    boo3 = sr_E3.get()
    if boo3 == "Y":
        ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
        ws.cell(row=84, column=9).value = ws.cell(row=84, column=9).value * req
        ws.cell(row=89, column=2).value = "Void of Flood"
        ws.cell(row=89, column=4).value = "yes"
        ws.cell(row=89, column=9).value = "yes"
    else: 
        ws.cell(row=89, column=2).value = ""
    boo4 = sr_E4.get()
    if boo4 == "Y":
        ws.cell(row=84, column=11).value = ws.cell(row=84, column=11).value * req
        ws.cell(row=84, column=19).value = ws.cell(row=84, column=19).value * req
        ws.cell(row=90, column=2).value = "Void of Urban Vulnerability"
        ws.cell(row=90, column=11).value = "yes"
        ws.cell(row=90, column=19).value = "yes"
    else: 
        ws.cell(row=90, column=2).value = ""
    boo5 = sr_E5.get()
    if boo5 == "Y":
        ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
        ws.cell(row=84, column=17).value = ws.cell(row=84, column=17).value * req
        ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
        ws.cell(row=91, column=2).value = "Void of Shock Absorbtion"
        ws.cell(row=91, column=4).value = "yes"
        ws.cell(row=91, column=17).value = "yes"
        ws.cell(row=91, column=7).value = "yes"
    else: 
        ws.cell(row=91, column=2).value = ""
    boo6 = sr_E6.get()
    if boo6 == "Y":
        ws.cell(row=87, column=2).value = "no voids present"
    wb.save(r'URSA Report.xlsx')

#Print DI
def printdi():
    boo7 = di_E1.get()
    if boo7 == "0":
        ws.cell(row=5, column=3).value = 0
    elif boo7 == "1":
      ws.cell(row=5, column=3).value = 1   
    elif boo7 == "2":
      ws.cell(row=5, column=3).value = 2
    elif boo7 == "3":
      ws.cell(row=5, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=5, column=i).value = ws.cell(row=5, column=i).value * ws.cell(row=5, column=3).value

    boo8 = di_E2.get()
    if boo8 == "0":
        ws.cell(row=9, column=3).value = 0
    elif boo8 == "1":
      ws.cell(row=9, column=3).value = 1   
    elif boo8 == "2":
      ws.cell(row=9, column=3).value = 2
    elif boo8 == "3":
      ws.cell(row=9, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=9, column=i).value = ws.cell(row=9, column=i).value * ws.cell(row=9, column=3).value

    boo9 = di_E3.get()
    if boo9 == "0":
        ws.cell(row=13, column=3).value = 0
    elif boo9 == "1":
      ws.cell(row=13, column=3).value = 1   
    elif boo9 == "2":
      ws.cell(row=13, column=3).value = 2
    elif boo9 == "3":
      ws.cell(row=13, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=13, column=i).value = ws.cell(row=13, column=i).value * ws.cell(row=13, column=3).value

    boo10 = di_E4.get()
    if boo10 == "0":
        ws.cell(row=19, column=3).value = 0
    elif boo10 == "1":
      ws.cell(row=19, column=3).value = 1   
    elif boo10 == "2":
      ws.cell(row=19, column=3).value = 2
    elif boo10 == "3":
      ws.cell(row=19, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=19, column=i).value = ws.cell(row=19, column=i).value * ws.cell(row=19, column=3).value

    boo11 = di_E5.get()
    if boo11 == "0":
        ws.cell(row=22, column=3).value = 0
    elif boo11 == "1":
      ws.cell(row=22, column=3).value = 1   
    elif boo11 == "2":
      ws.cell(row=22, column=3).value = 2
    elif boo11 == "3":
      ws.cell(row=22, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=22, column=i).value = ws.cell(row=22, column=i).value * ws.cell(row=22, column=3).value

    boo12 = di_E6.get()
    if boo12 == "0":
        ws.cell(row=24, column=3).value = 0
    elif boo12 == "1":
      ws.cell(row=24, column=3).value = 1   
    elif boo12 == "2":
      ws.cell(row=24, column=3).value = 2
    elif boo12 == "3":
      ws.cell(row=24, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=24, column=i).value = ws.cell(row=24, column=i).value * ws.cell(row=24, column=3).value

    boo13 = di_E7.get()
    if boo13 == "0":
        ws.cell(row=29, column=3).value = 0
    elif boo13 == "1":
      ws.cell(row=29, column=3).value = 1   
    elif boo13 == "2":
      ws.cell(row=29, column=3).value = 2
    elif boo13 == "3":
      ws.cell(row=29, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=29, column=i).value = ws.cell(row=29, column=i).value * ws.cell(row=29, column=3).value

    boo14 = di_E8.get()
    if boo14 == "0":
        ws.cell(row=34, column=3).value = 0
    elif boo14 == "1":
      ws.cell(row=34, column=3).value = 1   
    elif boo14 == "2":
      ws.cell(row=34, column=3).value = 2
    elif boo14 == "3":
      ws.cell(row=34, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=34, column=i).value = ws.cell(row=34, column=i).value * ws.cell(row=34, column=3).value

    boo15 = di_E9.get()
    if boo15 == "0":
        ws.cell(row=37, column=3).value = 0
    elif boo15 == "1":
      ws.cell(row=37, column=3).value = 1   
    elif boo15 == "2":
      ws.cell(row=37, column=3).value = 2
    elif boo15 == "3":
      ws.cell(row=37, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=37, column=i).value = ws.cell(row=37, column=i).value * ws.cell(row=37, column=3).value

    boo16 = di_E10.get()
    if boo16 == "0":
        ws.cell(row=42, column=3).value = 0
    elif boo16 == "1":
      ws.cell(row=42, column=3).value = 1   
    elif boo16 == "2":
      ws.cell(row=42, column=3).value = 2
    elif boo16 == "3":
      ws.cell(row=42, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=42, column=i).value = ws.cell(row=42, column=i).value * ws.cell(row=42, column=3).value

    boo17 = di_E11.get()
    if boo17 == "0":
        ws.cell(row=47, column=3).value = 0
    elif boo17 == "1":
      ws.cell(row=47, column=3).value = 1   
    elif boo17 == "2":
      ws.cell(row=47, column=3).value = 2
    elif boo17== "3":
      ws.cell(row=47, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=47, column=i).value = ws.cell(row=47, column=i).value * ws.cell(row=47, column=3).value

    boo18 = di_E12.get()
    if boo18 == "0":
        ws.cell(row=51, column=3).value = 0
    elif boo18 == "1":
      ws.cell(row=51, column=3).value = 1   
    elif boo18 == "2":
      ws.cell(row=51, column=3).value = 2
    elif boo18== "3":
      ws.cell(row=51, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=51, column=i).value = ws.cell(row=51, column=i).value * ws.cell(row=51, column=3).value

    boo19 = di_E13.get()
    if boo19 == "0":
        ws.cell(row=59, column=3).value = 0
    elif boo19 == "1":
      ws.cell(row=59, column=3).value = 1   
    elif boo19 == "2":
      ws.cell(row=59, column=3).value = 2
    elif boo19== "3":
      ws.cell(row=59, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=59, column=i).value = ws.cell(row=59, column=i).value * ws.cell(row=59, column=3).value

    boo20 = di_E14.get()
    if boo20 == "0":
        ws.cell(row=62, column=3).value = 0
    elif boo20 == "1":
      ws.cell(row=62, column=3).value = 1   
    elif boo20 == "2":
      ws.cell(row=62, column=3).value = 2
    elif boo20== "3":
      ws.cell(row=62, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=62, column=i).value = ws.cell(row=62, column=i).value * ws.cell(row=62, column=3).value

    boo21 = di_E15.get()
    if boo21 == "0":
        ws.cell(row=67, column=3).value = 0
    elif boo21 == "1":
      ws.cell(row=67, column=3).value = 1   
    elif boo21 == "2":
      ws.cell(row=67, column=3).value = 2
    elif boo21== "3":
      ws.cell(row=67, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=67, column=i).value = ws.cell(row=67, column=i).value * ws.cell(row=67, column=3).value

    boo22 = di_E16.get()
    if boo22 == "0":
        ws.cell(row=72, column=3).value = 0
    elif boo22== "1":
      ws.cell(row=72, column=3).value = 1   
    elif boo22 == "2":
      ws.cell(row=72, column=3).value = 2
    elif boo22== "3":
      ws.cell(row=72, column=3).value = 3  
    for i in range(4,19):
      ws.cell(row=72, column=i).value = ws.cell(row=72, column=i).value * ws.cell(row=72, column=3).value
    wb.save(r'URSA Report.xlsx')
    
#Print Development Impact
def printtotaldi():
    for i in range(4,20):
        ws.cell(row=83, column=i).value = ws.cell(row=5, column=i).value + ws.cell(row=9, column=i).value + ws.cell(row=13, column=i).value + ws.cell(row=19, column=i).value + ws.cell(row=22, column=i).value + ws.cell(row=24, column=i).value + ws.cell(row=29, column=i).value + ws.cell(row=34, column=i).value + ws.cell(row=37, column=i).value + ws.cell(row=42, column=i).value + ws.cell(row=47, column=i).value + ws.cell(row=51, column=i).value + ws.cell(row=59, column=i).value + ws.cell(row=62, column=i).value + ws.cell(row=67, column=i).value + ws.cell(row=72, column=i).value
        wb.save(r'URSA Report.xlsx')
printtotaldi()

#Print Difference
for i in range(4,20):
    global diff
    diff = ws.cell(row=84, column=i).value - ws.cell(row=83, column=i).value
    if ws.cell(row=83, column=i).value >= ws.cell(row=84, column=i).value:
        ws.cell(row=85, column=i).value = "sufficient"
    else:
        ws.cell(row=85, column=i).value = diff
for i in range(4,20):
    ws.cell(row=84, column=i).value
    wb.save(r'URSA Report.xlsx')

#URSA REPORT
#Extract from database
wb = Workbook ()
wb = openpyxl.load_workbook (r'URSA Report.xlsx', data_only = True)
ws = wb.worksheets[0]

#Function to open excel report
def open_excel ():
    os.startfile(r'URSA Report.xlsx')

#URSA REPORT -Text
Label(root, text='URSA Report', font=('Industry Inc Base', 12, 'normal')).place(x=950, y=150)
#Label(root, text='Developed by', font=('Myanmar Text', 7, 'normal')).place(x=950, y=585)
Label(root, text='Developed by Jerry Xinchen Yang, Grace Jing Yuan Yu and Karl Leung MSA, 2021.', font=('Myanmar Text', 7, 'normal')).place(x=955, y=600)


#Button print report to Excel
print_sr_inp1_button = Button(root,text="Print Inputs", width=40, bg='#2a2a2a',fg='#ffffff', font=('Industry Inc Base', 9, 'normal'), command=lambda: [printdn(),printsitereq(),printdi(),printtotaldi()]) 
print_sr_inp1_button.place(x=950, y=190)
#Label(root, text='Please view the entire URSA Report in your directory', fg='#a80f21',font=('Industry Inc Base', 9, 'normal')).place(x=1000, y=600)

#Button to open Excel
ursareport_button = Button(root,text="View URSA Report", width=40, height = 2, fg='#ffffff',font=('Industry Inc Base', 9, 'normal'),bg='#3C0947', command=open_excel) 
ursareport_button.place(x=950, y=250)

# # #text widget displaying URSA Report extract from Excel
# # text = Text (root, width=47, height=15, font=('Industry Inc Base', 9, 'normal') )
# # text.insert(INSERT, ws.cell(row=85, column=i).value)
# # text.insert(END, '\n')
# # text.place (x=1000, y=300)

# def printursareport():
#     for i in range(4,20): 
#         ws.cell(row=85, column=i).value

# printursareport = StringVar
# printursareport.get()

# label_ur = Label(root,text=printursareport.get())
# label_ur.place(x=950,y=500)

#ROOT WINDOW END
root.mainloop()