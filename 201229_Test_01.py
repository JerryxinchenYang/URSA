import openpyxl
from openpyxl import Workbook

wb =  openpyxl.load_workbook ('201227_DataSpread-JY.xlsx', data_only = True)
ws = wb.worksheets[0]

# print(ws.cell(row=3, column=4).value)

#user greeting
name = input("What's your name? ")
print("It's nice to meet you,", name)

print(name,',','The data base you will be using today is', wb)

# checking sheet names
# sn = wb.sheetnames
# print(sn)

print("Does your development exist within a void? If so, which?")
options = ["Void of Digital Connection", "Void of Urban Connectivity", "Void of Flood", "voids of urban Vulnerability", "Void of Shock Absorbtion","No Void"]

boo2 = ""
boo3 = ""
boo4 = ""
inp2 = ""
inp3 = ""
inp4 = ""
inp5 = ""
inp = ""
impact = ""
req = 1.3

# Print Site Requirement
def aerial_req():
    for i in range(4,20):
        ws.cell(row=84, column=i).value = ws.cell(row=5, column=i).value + ws.cell(row=9, column=i).value + ws.cell(row=13, column=i).value + ws.cell(row=19, column=i).value + ws.cell(row=22, column=i).value + ws.cell(row=24, column=i).value + ws.cell(row=29, column=i).value + ws.cell(row=31, column=i).value + ws.cell(row=34, column=i).value + ws.cell(row=37, column=i).value + ws.cell(row=42, column=i).value + ws.cell(row=47, column=i).value + ws.cell(row=51, column=i).value + ws.cell(row=59, column=i).value + ws.cell(row=62, column=i).value + ws.cell(row=67, column=i).value + ws.cell(row=72, column=i).value

aerial_req()

# defining functions needed for asking about user's site condition
def checkboo2():
    global boo2, inp3
    boo2 = input("Is there another void on site? Y for yes / N for no")
    if boo2 == "Y":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        inp3 = int(input("Enter a number: "))
        if inp3 == 1:
            print("you have picked", inp3, ", on top of your previous void(s), your development is required to have a higher impact in the 'Technology + information' category.")
            ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
        elif inp3 == 2:
            print("you have picked", inp3, ", on top of your previous void(s), your development is required to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
            ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
            ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
            ws.cell(row=84, column=6).value = ws.cell(row=84, column=6).value * req
        elif inp3 == 3:
            print("you have picked", inp3, ", on top of your previous void(s), your development is required to have a higher impact in the 'Water' and 'Green Infratsture' categories")
            ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
            ws.cell(row=84, column=9).value = ws.cell(row=84, column=9).value * req
        elif inp3 == 4:
            print("you have picked", inp3, ", on top of your previous void(s), your development is required to have a higher impact in the 'Building and Design'and'Economy'categories.")
            ws.cell(row=84, column=11).value = ws.cell(row=84, column=11).value * req
            ws.cell(row=84, column=19).value = ws.cell(row=84, column=19).value * req
        elif inp3 == 5:
            print("you have picked", inp3, ", on top of your previous void(s), your development is required to have a higher impact in the 'water','planning' and 'Connectivity'categories.")
            ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
            ws.cell(row=84, column=17).value = ws.cell(row=84, column=17).value * req
            ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
        else:
            print("Invalid input!")
    elif boo2 == "N":
        for i in range(len(options)):
                print(str(i+1) + ":", options[i])
        print("Your Final selection is", inp1, "and", inp2)
    else:
        print("Invalid input!")

def checkboo3():
    global boo3, inp3, inp4
    boo3 = input("Is there another void on site? Y for yes / N for no")
    if boo3 == "Y":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        inp4 = int(input("Enter a number: "))
        if inp4 == 1:
            print("you have picked", inp4, ", on top of your previous void(s), your development is required to have a higher impact in the 'Technology + information' category.")
            ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
        elif inp4 == 2:
            print("you have picked", inp4, ", on top of your previous void(s), your development is required to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
            ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
            ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
            ws.cell(row=84, column=6).value = ws.cell(row=84, column=6).value * req
        elif inp4 == 3:
            print("you have picked", inp4, ", on top of your previous void(s), your development is required to have a higher impact in the 'Water' and 'Green Infratsture' categories.")
            ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
            ws.cell(row=84, column=9).value = ws.cell(row=84, column=9).value * req
        elif inp4 == 4:
            print("you have picked", inp4, ", on top of your previous void(s), your development is required to have a higher impact in the 'Building and Design'and'Economy'categories.")
            ws.cell(row=84, column=11).value = ws.cell(row=84, column=11).value * req
            ws.cell(row=84, column=19).value = ws.cell(row=84, column=19).value * req
        elif inp4 == 5:
            print("you have picked", inp4, ", on top of your previous void(s), your development is required to have a higher impact in the 'water','planning' and 'Connectivity'categories.")
            ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
            ws.cell(row=84, column=17).value = ws.cell(row=84, column=17).value * req
            ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
        else:
            print("Invalid input!")
    elif boo3 == "N":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        print("Your Final selection is", inp1, "and", inp2, "and", inp3)
    else:
        print("Invalid input!")

def checkboo4():
    global boo4, inp3, inp4, inp5
    boo4 = input("Is there another void on site? Y for yes / N for no")
    if boo4 == "Y":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        inp5 = int(input("Enter a number: "))
        if inp5 == 1:
            print("you have picked", inp5, ", on top of your previous void(s), your development is required to have a higher impact in the 'Technology + information' category.")
            ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
        elif inp5 == 2:
            print("you have picked", inp5, ", on top of your previous void(s), your development is required to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
            ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
            ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
            ws.cell(row=84, column=6).value = ws.cell(row=84, column=6).value * req
        elif inp5 == 3:
            print("you have picked", inp5, ", on top of your previous void(s), your development is required to have a higher impact in the 'Water' and 'Green Infratsture' categories")
            ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
            ws.cell(row=84, column=9).value = ws.cell(row=84, column=9).value * req
        elif inp5 == 4:
            print("you have picked", inp5, ", on top of your previous void(s), your development is required to have a higher impact in the 'Building and Design'and'Economy'categories.")
            ws.cell(row=84, column=11).value = ws.cell(row=84, column=11).value * req
            ws.cell(row=84, column=19).value = ws.cell(row=84, column=19).value * req
        elif inp5 == 5:
            print("you have picked", inp5, ", on top of your previous void(s), your development is required to have a higher impact in the 'water','planning' and 'Connectivity'categories.")
            ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
            ws.cell(row=84, column=17).value = ws.cell(row=84, column=17).value * req
            ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
        else:
            print("Invalid input!")
        print("Your Final selection is", inp1, "and", inp2, "and", inp3,"and", inp4, "and", inp5)
    elif boo4 == "N":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        print("Your Final selection is", inp1, "and", inp2, "and", inp3,"and", inp4)
    else:
        print("Invalid input!")

# Print out options
for i in range(len(options)):
    print(str(i+1) + ":", options[i])

# Initial question on void condition within development
inp1 = int(input("Enter a number: "))
if inp1 == 1:
    print("you have picked", inp1, ", your development is required to have a higher impact in the 'Technology + information' category.")
    ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
elif inp1 == 2:
    print("you have picked", inp1, ", your development is required to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
    ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
    ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
    ws.cell(row=84, column=6).value = ws.cell(row=84, column=6).value * req
elif inp1 == 3:
    print("you have picked", inp1, ", your development is required to have a higher impact in the 'Water' and 'Green Infratsture' categories")
    ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
    ws.cell(row=84, column=9).value = ws.cell(row=84, column=9).value * req
elif inp1 == 4:
    print("you have picked", inp1, ",  your development is required to have a higher impact in the 'Building and Design'and'Economy'categories.")
    ws.cell(row=84, column=11).value = ws.cell(row=84, column=11).value * req
    ws.cell(row=84, column=19).value = ws.cell(row=84, column=19).value * req
elif inp1 == 5:
    print("you have picked", inp1, ", your development is required to have a higher impact in the 'water','planning' and 'Connectivity'categories.")
    ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
    ws.cell(row=84, column=17).value = ws.cell(row=84, column=17).value * req
    ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
elif inp1 == 6:
    print("you have picked", inp1, ", your development is not required to have a higher impact in a specific area .")
else:
    print("Invalid input!")


# Asking user if there is another void on site
boo1 = input("Is there another void on site? Y for yes / N for no")
if boo1 == "Y":
    for i in range(len(options)):
        print(str(i+1) + ":", options[i])
    inp2 = int(input("Enter a number: "))
    if inp2 == 1:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Technology + information' category.")
        ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
    elif inp2 == 2:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
        ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
        ws.cell(row=84, column=12).value = ws.cell(row=84, column=12).value * req
        ws.cell(row=84, column=6).value = ws.cell(row=84, column=6).value * req
    elif inp2 == 3:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Water' and 'Green Infratsture' categories")
        ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
        ws.cell(row=84, column=9).value = ws.cell(row=84, column=9).value * req
    elif inp2 == 4:
        print("you have picked", inp2, ", on top of your previous void(s),  your development is expected to have a higher impact in the 'Building and Design'and'Economy'categories.")
        ws.cell(row=84, column=11).value = ws.cell(row=84, column=11).value * req
        ws.cell(row=84, column=19).value = ws.cell(row=84, column=19).value * req
    elif inp2 == 5:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'water','planning' and 'Connectivity'categories.")
        ws.cell(row=84, column=4).value = ws.cell(row=84, column=4).value * req
        ws.cell(row=84, column=17).value = ws.cell(row=84, column=17).value * req
        ws.cell(row=84, column=7).value = ws.cell(row=84, column=7).value * req
elif boo1 == "N":
    for i in range(len(options)):
        print(str(i+1) + ":", options[i])
    print("Your Final selection is", inp1)
else:
    print("Invalid input!")

# Continue asking whether there are further voids on site
if boo1 != "N":
    checkboo2()
    if boo2 != "N":
        checkboo3()
        if boo3 != "N":
            checkboo4()

# print name and impact of voids selected into excel
def printinp(inp,cellnum):
    if inp == 1:
        ws.cell(row=cellnum, column=2).value = "Void of Digital Connection"
        ws.cell(row=cellnum, column=12).value = "yes"
    elif inp == 2:
        ws.cell(row=cellnum, column=2).value = "Void of Urban Connectivity"
        ws.cell(row=cellnum, column=12).value = "yes"
        ws.cell(row=cellnum, column=7).value = "yes"
        ws.cell(row=cellnum, column=6).value = "yes"
    elif inp == 3:
        ws.cell(row=cellnum, column=2).value = "Void of Flood"
        ws.cell(row=cellnum, column=4).value = "yes"
        ws.cell(row=cellnum, column=9).value = "yes"
    elif inp == 4:
        ws.cell(row=cellnum, column=2).value = "voids of urban Vulnerability"
        ws.cell(row=cellnum, column=11).value = "yes"
        ws.cell(row=cellnum, column=19).value = "yes"
    elif inp == 5:
        ws.cell(row=cellnum, column=2).value = "Void of Shock Absorbtion"
        ws.cell(row=cellnum, column=4).value = "yes"
        ws.cell(row=cellnum, column=17).value = "yes"
        ws.cell(row=cellnum, column=7).value = "yes"
    else:
        ws.cell(row=cellnum, column=2).value = "none"

printinp(inp1,87)
printinp(inp2,88)
printinp(inp3,89)
printinp(inp4,90)
printinp(inp5,91)

# Function asking user about their building impact
def askimpact(cellnum2):
    category = ws.cell(row=cellnum2, column=1).value
    print("The current category in question is", category)
    impact = int(input("On a scale of 0 to 3, how strongly is your development impacting this category?"))
    if impact == 0:
        print("Your development does not have any impact on the category of", category, ".")
        ws.cell(row=cellnum2, column=3).value = 0
    elif impact == 1:
        print("Your development has a small impact on the category of", category, ".")
        ws.cell(row=cellnum2, column=3).value = 1   
    elif impact == 2:
        print("Your development has a moderate impact on the category of", category, ".")
        ws.cell(row=cellnum2, column=3).value = 2
    elif impact == 3:
        print("Your development has a high impact on the category of", category, ".")
        ws.cell(row=cellnum2, column=3).value = 3
    else:
        print("Invalid Input!")

    for i in range(4,19):
        ws.cell(row=cellnum2, column=i).value = ws.cell(row=cellnum2, column=i).value * ws.cell(row=cellnum2, column=3).value

# Ask user about their building impact in all categories
askimpact(5)
askimpact(9)
askimpact(13)
askimpact(19)
askimpact(22)
askimpact(24)
askimpact(29)
askimpact(31)
askimpact(34)
askimpact(37)
askimpact(42)
askimpact(47)
askimpact(51)
askimpact(59)
askimpact(62)
askimpact(67)
askimpact(72)
print("your development impact input will be printed into column C in the excel report.")

# add together the total imapct value of the development
def total_impact():
    for i in range(4,20):
        ws.cell(row=83, column=i).value = ws.cell(row=5, column=i).value + ws.cell(row=9, column=i).value + ws.cell(row=13, column=i).value + ws.cell(row=19, column=i).value + ws.cell(row=22, column=i).value + ws.cell(row=24, column=i).value + ws.cell(row=29, column=i).value + ws.cell(row=31, column=i).value + ws.cell(row=34, column=i).value + ws.cell(row=37, column=i).value + ws.cell(row=42, column=i).value + ws.cell(row=47, column=i).value + ws.cell(row=51, column=i).value + ws.cell(row=59, column=i).value + ws.cell(row=62, column=i).value + ws.cell(row=67, column=i).value + ws.cell(row=72, column=i).value

total_impact()
print("the overall impact of your development in each category, will be printed into row 83 in the excel report.")

# Compare total impact to aerial need

for i in range(4,20):
    global diff
    diff = ws.cell(row=84, column=i).value - ws.cell(row=83, column=i).value
    if ws.cell(row=83, column=i).value >= ws.cell(row=84, column=i).value:
        print (name, ",your development impact in the category of", ws.cell(row=3, column=i).value, "is sufficient.")
        ws.cell(row=85, column=i).value = "sufficient"
    else:
        print (name,",your development impact in the category of", ws.cell(row=3, column=i).value, "is insufficient, you will require to add an additional of", diff)
        ws.cell(row=85, column=i).value = diff

for i in range(4,20):
    ws.cell(row=84, column=i).value

print ("Thank you for using the interface, results will be printed into the excel file.")

# get the percentage of each impact category to derive impact priorities

# for i in range(4,20):
#     ws.cell(row=83, column=21) = 0 +  ws.cell(row=83, column=i).value

wb.save('210106_testresult.xlsx')

# Asking user if there is another void on site

    
# Asking user if there is another void on site


# sheet = wb['Sheet1']

# sheet['D5'].value 

# print(sheet['D5'].value)

#for i in range(4,83):print(sheet.cell(row=i,column=4).value)
# print(sheet.max_row)

# wb.create_sheet(title='My Sheet Name', index=1)

# wb.save ('example.xlsx')
